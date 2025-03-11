from nba_api.stats.endpoints import scoreboardv2, leaguegamefinder
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import json
import os
from nba_api.stats.static import teams
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

# Elo Rating Constants
K_FACTOR = 20         # Sensitivity of Elo adjustments
HOME_ADVANTAGE = 100    # Home team Elo boost
ELO_FILE = "elo_ratings.json"  # File to persist Elo ratings
EXCEL_FILE = "nba_elo_tracker.xlsx"  # Excel file to store daily Elo and predictions

# Fetch all NBA teams and create mappings
nba_teams = teams.get_teams()
NBA_TEAM_IDS = {team["id"] for team in nba_teams}  # Set of official NBA team IDs
NBA_TEAM_NAMES = {team["full_name"] for team in nba_teams}  # Set of official NBA team names
TEAM_ID_MAP = {team["id"]: team["full_name"] for team in nba_teams}
TEAM_ABBREV_MAP = {team["abbreviation"]: team["full_name"] for team in nba_teams}

def expected_win_prob(r_team, r_opponent):
    return 1 / (1 + 10 ** ((r_opponent - r_team) / 400))

def update_elo(r_team, r_opponent, outcome):
    expected = expected_win_prob(r_team, r_opponent)
    change = K_FACTOR * (outcome - expected)
    return r_team + change

def predict_winner(home_team, away_team, elo_ratings):
    home_elo = elo_ratings.get(home_team, 1500) + HOME_ADVANTAGE
    away_elo = elo_ratings.get(away_team, 1500)
    home_win_prob = expected_win_prob(home_elo, away_elo)
    return home_team if home_win_prob > 0.5 else away_team, home_win_prob

def load_elo_ratings():
    if os.path.exists(ELO_FILE):
        with open(ELO_FILE, "r") as f:
            return json.load(f)
    return {}

def save_elo_ratings(elo_ratings):
    with open(ELO_FILE, "w") as f:
        json.dump(elo_ratings, f)

def is_nba_team(team_name):
    """Check if the team is an official NBA team."""
    return team_name in NBA_TEAM_NAMES

def initialize_elo_ratings():
    print("Initializing Elo ratings from season history...")
    elo_ratings = load_elo_ratings()
    if not elo_ratings:
        # Initialize only with official NBA teams
        elo_ratings = {team["full_name"]: 1500 for team in nba_teams}
    else:
        # Filter out non-NBA teams from existing ratings
        elo_ratings = {team: rating for team, rating in elo_ratings.items() if is_nba_team(team)}
    
    # Fetch season history (using last season as reference)
    season_str = f"{datetime.today().year - 1}-{str(datetime.today().year)[-2:]}"
    game_finder = leaguegamefinder.LeagueGameFinder(season_nullable=season_str)
    games = game_finder.get_data_frames()[0]
    
    # Group by GAME_ID to pair home and away rows for each game
    grouped = games.groupby("GAME_ID")
    for game_id, group in grouped:
        if len(group) < 2:
            continue  # Skip incomplete games
        
        # Identify home and away rows based on the MATCHUP string
        home_game = group[group["MATCHUP"].str.contains("vs.")]
        away_game = group[group["MATCHUP"].str.contains("@")]
        
        if home_game.empty or away_game.empty:
            continue
        
        home_row = home_game.iloc[0]
        away_row = away_game.iloc[0]
        
        home_team = home_row["TEAM_NAME"]
        away_team = away_row["TEAM_NAME"]
        
        # Skip non-NBA teams
        if not is_nba_team(home_team) or not is_nba_team(away_team):
            continue
            
        home_score = home_row["PTS"]
        away_score = away_row["PTS"]
        
        if pd.isna(home_score) or pd.isna(away_score):
            continue
        
        outcome = 1 if home_score > away_score else 0
        
        home_team_rating = elo_ratings.get(home_team, 1500) + HOME_ADVANTAGE
        away_team_rating = elo_ratings.get(away_team, 1500)
        
        new_home_rating = update_elo(home_team_rating, away_team_rating, outcome)
        new_away_rating = update_elo(away_team_rating, home_team_rating, 1 - outcome)
        
        # Remove home advantage after update for storage
        elo_ratings[home_team] = round(new_home_rating - HOME_ADVANTAGE, 2)
        elo_ratings[away_team] = round(new_away_rating, 2)
        
        print(f"Updated Elo for game {game_id}: {home_team} ({elo_ratings[home_team]}) vs {away_team} ({elo_ratings[away_team]})")
    
    save_elo_ratings(elo_ratings)
    return elo_ratings

def fetch_nba_games(start_date, end_date):
    try:
        games = []
        date_cursor = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_obj = datetime.strptime(end_date, "%Y-%m-%d")
        
        while date_cursor <= end_date_obj:
            formatted_date = date_cursor.strftime("%Y-%m-%d")
            scoreboard_data = scoreboardv2.ScoreboardV2(game_date=formatted_date).get_dict()
            game_result_sets = scoreboard_data.get("resultSets", [])
            
            if game_result_sets:
                game_data = game_result_sets[0]["rowSet"]
                for game in game_data:
                    home_team_id = game[6]
                    away_team_id = game[7]
                    
                    # Only include games between NBA teams
                    if home_team_id in NBA_TEAM_IDS and away_team_id in NBA_TEAM_IDS:
                        home_team_name = TEAM_ID_MAP.get(home_team_id, "Unknown")
                        away_team_name = TEAM_ID_MAP.get(away_team_id, "Unknown")
                        games.append({"date": formatted_date, "home_team": home_team_name, "away_team": away_team_name})
            
            date_cursor += timedelta(days=1)
        
        return games
    except Exception as e:
        print("Error fetching NBA games:", e)
        return []

def process_games(games, elo_ratings):
    predictions_by_day = {}
    for game in games:
        date = game['date']
        home_team = game['home_team']
        away_team = game['away_team']
        
        pred_winner, home_win_prob = predict_winner(home_team, away_team, elo_ratings)
        
        if date not in predictions_by_day:
            predictions_by_day[date] = []
        
        # Include Elo ratings in the prediction data
        home_elo = elo_ratings.get(home_team, 1500)
        away_elo = elo_ratings.get(away_team, 1500)
        
        # Store raw probability without rounding
        predictions_by_day[date].append([
            home_team, 
            away_team, 
            home_elo, 
            away_elo, 
            pred_winner, 
            home_win_prob  # Keep full precision
        ])
    
    return predictions_by_day

def apply_table_styles(sheet, start_row, end_row, start_col, end_col, is_header=False):
    """Apply styling to a table range in the worksheet."""
    # Define styles
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    alt_row_fill = PatternFill(start_color="E6F0FF", end_color="E6F0FF", fill_type="solid")
    border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # Apply styles to the range
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            
            # Apply borders to all cells
            cell.border = border
            
            # Apply header styling
            if is_header or row == start_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # Apply alternating row colors
                if (row - start_row) % 2 == 1:
                    cell.fill = alt_row_fill
                
                # Center alignment for all cells
                cell.alignment = Alignment(horizontal="center", vertical="center")

def save_to_excel(elo_ratings, predictions_by_day):
    """Save Elo ratings and predictions to Excel with date-based sheets."""
    # Create workbook if it doesn't exist
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    else:
        workbook = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]
    
    for date, predictions in predictions_by_day.items():
        # Check if sheet for this date already exists
        if date in workbook.sheetnames:
            print(f"Sheet for {date} already exists. Skipping...")
            continue
        
        # Create new sheet for this date
        sheet = workbook.create_sheet(title=date)
        
        # Add title and format date
        date_obj = datetime.strptime(date, "%Y-%m-%d")
        formatted_date = date_obj.strftime("%B %d, %Y")
        sheet.cell(row=1, column=1, value=f"NBA Elo Ratings and Predictions - {formatted_date}")
        sheet.cell(row=1, column=1).font = Font(size=16, bold=True)
        sheet.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
        sheet.row_dimensions[1].height = 30
        
        # Add space
        sheet.row_dimensions[2].height = 10
        
        # Add Elo ratings table
        ratings_header_row = 3
        sheet.cell(row=ratings_header_row, column=1, value="NBA TEAM ELO RATINGS")
        sheet.cell(row=ratings_header_row, column=1).font = Font(size=14, bold=True)
        sheet.merge_cells(start_row=ratings_header_row, end_row=ratings_header_row, start_column=1, end_column=3)
        
        # Ratings table headers
        sheet.cell(row=ratings_header_row+1, column=1, value="Rank")
        sheet.cell(row=ratings_header_row+1, column=2, value="Team")
        sheet.cell(row=ratings_header_row+1, column=3, value="Elo Rating")
        
        # Sort teams by Elo rating
        sorted_elos = sorted(
            [(team, rating) for team, rating in elo_ratings.items() if is_nba_team(team)],
            key=lambda x: x[1],
            reverse=True
        )
        
        # Add team ratings
        for i, (team, rating) in enumerate(sorted_elos):
            row = ratings_header_row + 2 + i
            sheet.cell(row=row, column=1, value=i+1)  # Rank
            sheet.cell(row=row, column=2, value=team)
            sheet.cell(row=row, column=3, value=rating)
        
        # Style the ratings table
        apply_table_styles(
            sheet,
            ratings_header_row+1,
            ratings_header_row+1+len(sorted_elos),
            1,
            3,
            is_header=True
        )
        
        # Adjust column widths
        sheet.column_dimensions["A"].width = 10
        sheet.column_dimensions["B"].width = 30
        sheet.column_dimensions["C"].width = 15
        
        # Add space between tables
        predictions_header_row = ratings_header_row + len(sorted_elos) + 4
        sheet.row_dimensions[predictions_header_row-1].height = 20
        
        # Add predictions table
        sheet.cell(row=predictions_header_row, column=1, value="GAME PREDICTIONS")
        sheet.cell(row=predictions_header_row, column=1).font = Font(size=14, bold=True)
        sheet.merge_cells(start_row=predictions_header_row, end_row=predictions_header_row, start_column=1, end_column=6)
        
        # Predictions table headers
        sheet.cell(row=predictions_header_row+1, column=1, value="Home Team")
        sheet.cell(row=predictions_header_row+1, column=2, value="Away Team")
        sheet.cell(row=predictions_header_row+1, column=3, value="Home Elo")
        sheet.cell(row=predictions_header_row+1, column=4, value="Away Elo")
        sheet.cell(row=predictions_header_row+1, column=5, value="Predicted Winner")
        sheet.cell(row=predictions_header_row+1, column=6, value="Home Team Win Probability")  # Renamed for clarity
        
        # Add prediction data
        for i, pred in enumerate(predictions):
            row = predictions_header_row + 2 + i
            home_team, away_team, home_elo, away_elo, winner, probability = pred
            
            sheet.cell(row=row, column=1, value=home_team)
            sheet.cell(row=row, column=2, value=away_team)
            sheet.cell(row=row, column=3, value=home_elo)
            sheet.cell(row=row, column=4, value=away_elo)
            sheet.cell(row=row, column=5, value=winner)
            sheet.cell(row=row, column=6, value=probability)
            
            # Format the probability as percentage with 3 decimal places
            sheet.cell(row=row, column=6).number_format = '0.000'
        
        # Style the predictions table
        apply_table_styles(
            sheet,
            predictions_header_row+1,
            predictions_header_row+1+len(predictions),
            1,
            6,
            is_header=True
        )
        
        # Adjust column widths for prediction table
        sheet.column_dimensions["D"].width = 15
        sheet.column_dimensions["E"].width = 30
        sheet.column_dimensions["F"].width = 30
        
        # Add conditional formatting for the win probability
        # Probabilities over 0.5 (favoring home team) will have green background
        # Probabilities under 0.5 (favoring away team) will have pink background
        prob_start_row = predictions_header_row + 2
        prob_end_row = predictions_header_row + 1 + len(predictions)
        
        # Define conditional formatting rule for home team favored
        sheet.conditional_formatting.add(
            f"F{prob_start_row}:F{prob_end_row}",
            Rule(
                type="cellIs",
                operator="greaterThan",
                formula=["0.5"],
                stopIfTrue=True,
                dxf=DifferentialStyle(fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
            )
        )
        
        # Define conditional formatting rule for away team favored
        sheet.conditional_formatting.add(
            f"F{prob_start_row}:F{prob_end_row}",
            Rule(
                type="cellIs",
                operator="lessThan",
                formula=["0.5"],
                stopIfTrue=True,
                dxf=DifferentialStyle(fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
            )
        )
    
    # Save the workbook
    workbook.save(EXCEL_FILE)
    print(f"Data saved to {EXCEL_FILE}")

def predict_games_for_range(start_date, end_date):
    print(f"Predicting games from {start_date} to {end_date}...\n")
    
    elo_ratings = initialize_elo_ratings()
    
    print("\nCurrent Elo Ratings:")
    # Filter out non-NBA teams before displaying
    nba_elo_ratings = {team: rating for team, rating in elo_ratings.items() if is_nba_team(team)}
    elo_df = pd.DataFrame(nba_elo_ratings.items(), columns=["Team", "Elo Rating"]).sort_values(by="Elo Rating", ascending=False)
    print(elo_df.to_string(index=False))
    
    games = fetch_nba_games(start_date, end_date)
    predictions_by_day = process_games(games, elo_ratings)
    save_elo_ratings(elo_ratings)
    
    # Save to Excel
    save_to_excel(elo_ratings, predictions_by_day)
    
    print("\nPredictions for the Selected Date Range:")
    for date, games in predictions_by_day.items():
        print(f"\n📅 **{date}**")
        df = pd.DataFrame(games, columns=['Home_Team', 'Away_Team', 'Home_Elo', 'Away_Elo', 'Predicted_Winner', 'Home_Team_Win_Probability'])
        print(df.to_string(index=False))

if __name__ == "__main__":
    # Predict games for today by default
    today = datetime.today().strftime("%Y-%m-%d")
    predict_games_for_range(today, today)
